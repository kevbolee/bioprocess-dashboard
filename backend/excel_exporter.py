"""
Excel exporters for analytical results.

Produces .xlsx workbooks that match the column layout users expect (see
excelexport/dataexample.xlsx):
  - vicell   → single sheet "ViCell_data"    (16 columns; A-C are formulas)
  - nova     → single sheet "Flex_2_Data"    (56 columns; A-C are formulas)
  - cedex    → single sheet "Cedex_Data"     (61 columns; A-D are formulas)

The first 3-4 columns are Excel formulas that parse the sample identifier
into a splice ID, sample type, and dilution factor. The formulas are written
per row referencing the correct row number so they auto-calculate on open.

For fields we do not store in our local database (e.g. Cedex Raw01-Raw40,
ViCell "Average cells / image"), the cell is left blank — the formulas still
compute off the sample identifier and users can paste additional data by hand
just like in the example workbook.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import database as db


_DATETIME_FMT = "yyyy-mm-dd hh:mm:ss"
_DATE_INT_FMT = "0"  # Cedex "Date" column is stored as YYYYMMDD integer
_TIME_FMT     = "hh:mm:ss"


def _parse_dt(s: Optional[str]) -> Optional[datetime]:
    """Best-effort parse of the various date/time strings we store."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    s = str(s).strip()
    if not s:
        return None
    # Common shapes: ISO (2026-04-21T09:51:48) or space-separated with optional TZ
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d"):
        try:
            return datetime.strptime(s.split("+")[0].split("Z")[0], fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _header_font() -> Font:
    return Font(bold=True)


def _write_headers(ws, headers: Iterable[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = _header_font()


def _autosize(ws, max_col: int, cap: int = 24) -> None:
    """Very light auto-fit: only widen a couple of columns that hold long labels."""
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        header_val = ws.cell(row=1, column=col_idx).value or ""
        ws.column_dimensions[letter].width = min(max(12, len(str(header_val)) + 2), cap)


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ViCell
# ---------------------------------------------------------------------------

_VICELL_HEADERS = [
    "Splice ID", "Dilution Factor", "sampletype", "sample_identifier", "epoch_time",
    "Viability (%)", "Average diameter (microns)", "Total cells", "Viable cells",
    "Cell type", "Dilution factor", "Total cells / ml (x10^6)",
    "Viable cells / ml (x10^6)", "Average circularity", "Average cells / image",
    "Average background intensity",
]


def build_vicell_xlsx(since: Optional[str] = None,
                      until: Optional[str] = None,
                      sample_id_filter: Optional[str] = None,
                      days_back: int = 3650) -> bytes:
    """Return a workbook matching the example's ViCell_data sheet."""
    rows = db.get_vicell_results(days_back=days_back,
                                 sample_id_filter=sample_id_filter,
                                 since=since, until=until)
    # Chronological (oldest → newest) reads more naturally in an export.
    rows.sort(key=lambda r: r.get("sample_date") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "ViCell_data"
    _write_headers(ws, _VICELL_HEADERS)

    for i, r in enumerate(rows, start=2):
        # Cols A-C: formulas that parse col D (sample_identifier)
        ws.cell(row=i, column=1, value=f'=LEFT(D{i},FIND("-",D{i},FIND("-",D{i})+1)-1)')
        ws.cell(row=i, column=2, value=f'=LEFT(RIGHT(D{i},LEN(D{i})-FIND("-",D{i},9)),2)')
        ws.cell(row=i, column=3, value=f"=A{i}&RIGHT(D{i},1)")

        ws.cell(row=i, column=4, value=r.get("sample_id"))

        dt = _parse_dt(r.get("sample_date"))
        if dt is not None:
            c = ws.cell(row=i, column=5, value=dt)
            c.number_format = _DATETIME_FMT
        else:
            ws.cell(row=i, column=5, value=r.get("sample_date"))

        ws.cell(row=i, column=6,  value=r.get("viability_pct"))
        ws.cell(row=i, column=7,  value=r.get("avg_diameter_um"))
        ws.cell(row=i, column=8,  value=r.get("total_cells"))
        ws.cell(row=i, column=9,  value=r.get("viable_cells"))
        ws.cell(row=i, column=10, value=r.get("cell_type"))
        ws.cell(row=i, column=11, value=r.get("dilution_factor"))
        ws.cell(row=i, column=12, value=r.get("total_cells_per_ml"))
        ws.cell(row=i, column=13, value=r.get("viable_cells_per_ml"))
        ws.cell(row=i, column=14, value=r.get("avg_circularity"))
        # Cols 15 (Average cells / image) and 16 (Average background intensity)
        # are not stored locally — leave blank.

    ws.freeze_panes = "A2"
    _autosize(ws, len(_VICELL_HEADERS))
    return _wb_bytes(wb)


# ---------------------------------------------------------------------------
# Nova Flex 2
# ---------------------------------------------------------------------------

_NOVA_HEADERS = [
    "Splice ID", "sampletype", "Dilution Factor", "sample_identifier", "epoch_time",
    "Sample Type", "Gln", "Glu", "Gluc", "Lac", "NH4+", "Na+", "K+", "Ca++",
    "pH", "PO2", "PCO2", "Osm", "Vessel ID", "Batch ID", "Cell Type",
    "Vessel Temperature", "Vessel Pressure", "Sparging O2%",
    "Chemistry Dilution Ratio", "pH @ Temp", "PCO2 @ Temp", "PO2 @ Temp",
    "O2 Saturation", "CO2 Saturation", "HCO3", "Chemistry Flow Time",
    "pH / Gas Flow Time", "Tray Location", "Time In Tray", "Sample Time",
    "Operator", "Total Density", "Viable Density", "Viability",
    "Avg. Live Diameter", "Total Live Count", "Total Cell Count",
    "Live Cell Std. Deviation", "Cell Inspection Type",
    "Pre-Dilution Multiplier", "Cell Density Dilution", "Valid Images",
    "Cell Density Flow", "Retain(s)", "Comment",
    "Chemistry Cartridge Lot Number", "Chemistry Card Lot Number",
    "Gas Cartridge Lot Number", "Gas Card Lot Number", "Exposure Time (msec)",
]

# Map Excel column (1-indexed) → (group_name, analyte) key used in nova_results.
# Only columns we actually collect from the Nova CSV are listed here; the rest
# stay blank in the export.
_NOVA_ANALYTE_MAP: dict[int, tuple[str, str]] = {
    7:  ("Chem",              "Gln"),
    8:  ("Chem",              "Glu"),
    9:  ("Chem",              "Gluc"),
    10: ("Chem",              "Lac"),
    11: ("Chem",              "NH4"),
    12: ("Chem",              "Na"),
    13: ("Chem",              "K"),
    14: ("Chem",              "Ca"),
    15: ("Gas",               "pH"),
    16: ("Gas",               "pO2"),
    17: ("Gas",               "pCO2"),
    18: ("Osmo",              "Osmolality"),
    26: ("CalculatedResults", "pHCorrected"),
    27: ("CalculatedResults", "pCO2Corrected"),
    28: ("CalculatedResults", "pO2Corrected"),
    29: ("CalculatedResults", "O2Saturation"),
    30: ("CalculatedResults", "CO2Saturation"),
    31: ("CalculatedResults", "HCO3"),
    38: ("CellDensity",       "TotalDensity"),
    39: ("CellDensity",       "ViableDensity"),
    40: ("CellDensity",       "Viability"),
    41: ("CellDensity",       "AvgLiveDiameter"),
    42: ("CellDensity",       "TotalLiveCount"),
    43: ("CellDensity",       "TotalCellCount"),
}


def build_nova_xlsx(since: Optional[str] = None,
                    until: Optional[str] = None,
                    days_back: int = 3650) -> bytes:
    """Return a workbook matching the example's Flex_2_Data sheet.

    Pivots our per-analyte nova_results rows back to one row per sample_time.
    """
    flat = db.get_nova_results(days_back=days_back, since=since, until=until)

    # Group by sample_time.
    grouped: dict[str, dict] = {}
    for r in flat:
        st = r.get("sample_time") or ""
        g = grouped.setdefault(st, {
            "sample_time": st,
            "sample_id":   r.get("sample_id"),
            "values":      {},
        })
        g["values"][(r.get("group_name"), r.get("analyte"))] = r.get("result_value")

    # Oldest → newest for a readable export.
    samples = sorted(grouped.values(), key=lambda s: s["sample_time"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Flex_2_Data"
    _write_headers(ws, _NOVA_HEADERS)

    for i, s in enumerate(samples, start=2):
        ws.cell(row=i, column=1, value=f'=LEFT(D{i},FIND("-",D{i},FIND("-",D{i})+1)-1)')
        ws.cell(row=i, column=2, value=f"=A{i}&RIGHT(D{i},1)")
        ws.cell(row=i, column=3, value=f'=LEFT(RIGHT(D{i},LEN(D{i})-FIND("-",D{i},9)),1)')

        ws.cell(row=i, column=4, value=s["sample_id"])

        dt = _parse_dt(s["sample_time"])
        if dt is not None:
            c = ws.cell(row=i, column=5, value=dt)
            c.number_format = _DATETIME_FMT
            c2 = ws.cell(row=i, column=36, value=dt)  # "Sample Time"
            c2.number_format = _DATETIME_FMT
        else:
            ws.cell(row=i, column=5,  value=s["sample_time"])
            ws.cell(row=i, column=36, value=s["sample_time"])

        for col_idx, key in _NOVA_ANALYTE_MAP.items():
            v = s["values"].get(key)
            if v is not None:
                ws.cell(row=i, column=col_idx, value=v)

    ws.freeze_panes = "A2"
    _autosize(ws, len(_NOVA_HEADERS))
    return _wb_bytes(wb)


# ---------------------------------------------------------------------------
# Cedex BIO HT
# ---------------------------------------------------------------------------

_CEDEX_HEADERS = [
    "Spliced ID", "FULLID", "Dilution Factor", "Sampletype",
    "Instr", "Msg", "Index", "App", "Date", "Time", "SW-Version", "Serial",
    "Test", "User", "Sample", "Order-Time", "Result", "Unit", "Flags", "Rate",
] + [f"Raw{n:02d}" for n in range(1, 41)] + ["IsSTAT"]


def build_cedex_xlsx(since: Optional[str] = None,
                     until: Optional[str] = None,
                     days_back: int = 3650) -> bytes:
    """Return a workbook matching the example's Cedex_Data sheet.

    Uses local bioht_results (imported from CEDEX BIO HT archive TXT files) —
    that's the same source as the columns in the example. Raw01-Raw40 and a
    few device-metadata columns are left blank since we don't currently store
    them.
    """
    rows = db.get_bioht_results(days_back=days_back, since=since, until=until)
    rows.sort(key=lambda r: r.get("sample_time") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cedex_Data"
    _write_headers(ws, _CEDEX_HEADERS)

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=f'=LEFT(O{i},FIND("-",O{i},FIND("-",O{i})+1)-1)')
        ws.cell(row=i, column=2, value=f"=A{i}&D{i}&LEFT(M{i},3)")
        ws.cell(row=i, column=3, value=f'=LEFT(RIGHT(O{i},LEN(O{i})-FIND("-",O{i},9)),1)')
        ws.cell(row=i, column=4, value=f'=RIGHT(O{i},1)')

        dt = _parse_dt(r.get("sample_time"))
        if dt is not None:
            # Col 9 "Date" = YYYYMMDD as int (matches example)
            date_int = dt.year * 10000 + dt.month * 100 + dt.day
            c9 = ws.cell(row=i, column=9, value=date_int)
            c9.number_format = _DATE_INT_FMT
            # Col 10 "Time"
            c10 = ws.cell(row=i, column=10, value=dt.time())
            c10.number_format = _TIME_FMT
        else:
            ws.cell(row=i, column=9,  value=r.get("sample_time"))

        ws.cell(row=i, column=13, value=r.get("test_abbrev"))          # Test
        ws.cell(row=i, column=15, value=r.get("sample_id"))            # Sample

        # Result: prefer numeric result_value; fall back to result_text
        result_val = r.get("result_value")
        if result_val is None:
            result_val = r.get("result_text")
        ws.cell(row=i, column=17, value=result_val)

        ws.cell(row=i, column=18, value=r.get("unit"))                 # Unit
        ws.cell(row=i, column=19, value=r.get("status"))               # Flags

        # Cols 5-8 (Instr/Msg/Index/App), 11-12 (SW/Serial), 14 (User), 16 (Order-Time),
        # 20 (Rate), 21-60 (Raw01-40), 61 (IsSTAT) — not stored, left blank.

    ws.freeze_panes = "A2"
    _autosize(ws, len(_CEDEX_HEADERS))
    return _wb_bytes(wb)
